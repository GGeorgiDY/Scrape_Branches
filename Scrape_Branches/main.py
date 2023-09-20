import time
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
from sent_email import sent_email
import pandas as pd


def scrape_fibank_branches():
    url = "https://my.fibank.bg/EBank/public/offices"
    driver = webdriver.Chrome()   # Initializing the web driver
    driver.get(url)   # Opening the webpage
    time.sleep(5)   # Waiting for the page to load
    page_source = driver.page_source   # Extracting the page source
    driver.quit()   # Closing the web driver
    soup = BeautifulSoup(page_source, "html.parser")   # Parsing the HTML using BeautifulSoup

    office_elements = soup.find_all("div", class_="sg-office-badge-view ng-scope")   # Find all bank offices

    # Create lists to store the extracted data
    office_names = []
    addresses = []
    telephones = []
    working_hours_saturday = []
    working_hours_sunday = []

    # Iterate through the office elements and extract the required information
    for office in office_elements:
        if not office.find("p", class_="ng-binding ng-scope", attrs={"bo-if": "worktime.satStartTime"}):
            continue

        office_name = office.find("p", attrs={"bo-bind": "item.name"}).text.strip()
        address = office.find("p", attrs={"bo-bind": "item.address"}).text.strip()
        telephone = office.find("p", class_="info-wrapper s2").text.strip()
        sat_hours = " ".join((office.find("p", class_="ng-binding ng-scope", attrs={"bo-if": "worktime.satStartTime"}).text.strip()).split()) if office.find("p", class_="ng-binding ng-scope", attrs={"bo-if": "worktime.satStartTime"}) else ""
        sun_hours = " ".join((office.find("p", class_="ng-binding ng-scope", attrs={"bo-if": "worktime.sunStartTime"}).text.strip()).split()) if office.find("p", class_="ng-binding ng-scope", attrs={"bo-if": "worktime.sunStartTime"}) else ""

        office_names.append(office_name)
        addresses.append(address)
        telephones.append(telephone)
        working_hours_saturday.append(sat_hours)
        working_hours_sunday.append(sun_hours)

    # Storing data using Pandas
    data = {
        "Office name": office_names,
        "Address": addresses,
        "Telephone": telephones,
        "Working hours Saturday": working_hours_saturday,
        "Working hours Sunday": working_hours_sunday,
    }
    df = pd.DataFrame(data)

    excel_file_path = r'C:\PythonApp\fibank_branches.xlsx'
    df.to_excel(excel_file_path, index=False, engine='openpyxl')

    # If I want to save the data in the Excel file without using pandas I can do it this way:
    # workbook = openpyxl.Workbook()
    # worksheet = workbook.active
    # worksheet.title = "Fibank Branches"
    #
    # headers = ["Office name", "Address", "Telephone", "Working hours Saturday", "Working hours Sunday"]
    #
    # for col_num, header in enumerate(headers, 1):
    #     worksheet.cell(row=1, column=col_num, value=header)
    #
    # for row_num, data_row in enumerate(zip(office_names, addresses, telephones, working_hours_saturday, working_hours_sunday), 2):
    #     for col_num, value in enumerate(data_row, 1):
    #         worksheet.cell(row=row_num, column=col_num, value=value)
    #
    # excel_file_path = r'C:\PythonApp\fibank_branches.xlsx'
    # workbook.save(excel_file_path)

    sent_email(excel_file_path)


if __name__ == "__main__":
    scrape_fibank_branches()


# necessary installments
# selenium
# beautifulsoup4
# pandas
# openpyxl
